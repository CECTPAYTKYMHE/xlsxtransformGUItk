import tkinter as tk
import tkinter.filedialog as fd
from tkinter import ttk
from typing import Sized
import openpyxl
from openpyxl_image_loader import SheetImageLoader
import os
import csv
import io
import glob
from PIL import Image
from tkinter.ttk import Progressbar
from tkinter import messagebox
import pandas as pd
import webbrowser

root = tk.Tk()
root.title('Работа с файлами для карт СКУД')
root.geometry('500x400+700+300')
root.resizable(False,False)


def export_emloyee():
    cwd = get_directory()
    os.mkdir(f'{cwd}/photo')
    os.mkdir(f'{cwd}/готовое')
    files = [f for f in glob.glob(cwd + '**/*.xlsx', recursive=True)]
    bars = 100 / len(files)
    bar = Progressbar(root, length=200, style='black.Horizontal.TProgressbar')
    bar.grid(row=7,column=1,stick='we')
    for file in files:
        bar['value'] += bars
        try:
            # filein = cwd + '/' + file
            wb = openpyxl.load_workbook(file)
            ws = wb.worksheets[0]
            i = 2 #начальная строка
            row = ws.max_row
            image_loader = SheetImageLoader(ws)
            ws.cell(row=1, column=5).value = 'Фотография №'
            ws.cell(row=1, column=6).value = 'Role'
            ws.cell(row=1, column=7).value = ''
            while i != row + 1: #проверить конечную строку!!!!
                ws.cell(row=i, column=6).value = 'Employee'
                ID = ws.cell(row=i, column=1)
                ID = str(ID.value)
                ID = ID.replace('-','')
                ID = ID.upper()
                try:
                    image = image_loader.get('H' + str(i)) #клетка с фотографией
                    if image.format.lower() in ['jpg', 'jpeg']:
                        image.save(f'{cwd}/photo/{ID}.jpg') #экспорт фото
                        ws.cell(row=i, column=5).value = ID + '.jpg' #имяфото.jpg
                    else:
                        logs = open(f'{cwd}/errorlogs.csv','a')
                        print(f'<Неправильная фотография в файле {file} в строке {i} у сотрудника {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value}>')
                        logs.write(f'{file}, {i}, {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value} \n')
                        logs.close()
                        pass
                except:
                    logs = open(f'{cwd}/errorlogs.csv','a')
                    print(f'<Неправильная фотография в файле {file} в строке {i} у сотрудника {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value}>')
                    logs.write(f'{file}, {i}, {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value} \n')
                    logs.close()
                    pass
                # ws.cell(row=i, column=5).value = ID +'.jpg' #имяфото.jpg
                ws.cell(row=i, column=1).value = ID #удаление '-' в NCFUGUID
                i += 1
            ws._images = [] # удаление всех фото из файла
            ws.delete_cols(7, 2)
            filename = file.split("\\")[-1]
            wb.save(filename = f'{cwd}/готовое/{filename}')
            wb.close()
        except:
            messagebox.showerror(title='Ошибка', message='Что-то пошло не так')
    messagebox.showinfo(title='Успех', message='Файлы готовы')



def export_student():
    cwd = get_directory()
    files = [f for f in glob.glob(cwd + '**/*.xlsx', recursive=True)]
    Image.MAX_IMAGE_PIXELS = None
    os.mkdir(f'{cwd}/photo')
    os.mkdir(f'{cwd}/готовое')
    bars = 100 / len(files)
    bar = Progressbar(root, length=200, style='black.Horizontal.TProgressbar')
    bar.grid(row=7,column=1,stick='we')
    try:
        logs = open(f'{cwd}/errorlogs.csv', 'x')
        logs.write('Ошибка фотографии в файле, Строка, ФИО\n')
        logs.close()
    except:
        pass
    for file in files:
        bar['value'] += bars
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.worksheets[0]
            i = 2 #начальная строка
            row = ws.max_row
            SheetImageLoader._images.clear()
            image_loader = SheetImageLoader(ws)
            ws.cell(row=1, column=1).value = 'ID'
            ws.cell(row=1, column=2).value = 'Фамилия'
            ws.cell(row=1, column=3).value = 'Имя'
            ws.cell(row=1, column=4).value = 'Отчество'
            ws.cell(row=1, column=5).value = 'Институт'
            ws.cell(row=1, column=6).value = 'Курс'
            ws.cell(row=1, column=7).value = 'Фотография №'
            while i != row + 1: #проверить конечную строку!!!!
                ID = ws.cell(row=i, column=1)
                ID = str(ID.value)
                ID = ID.replace('-','')
                ID = ID.upper()
                try:
                    ws.unmerge_cells(start_row=i, start_column=7, end_row=i, end_column=8)
                except:
                    pass
                try:
                    image = image_loader.get('G' + str(i)) #клетка с фотографией
                    if image.format.lower() in ['jpg', 'jpeg']:
                        image.save(f'{cwd}/photo/{ID}.jpg') #экспорт фото
                        ws.cell(row=i, column=7).value = ID + '.jpg' #имяфото.jpg
                    else:
                        print(f'<Неправильная фотография в файле {file} в строке {i-1} у студента {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value}>')
                        logs.write(f'{file}, {i-1}, {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value} \n')
                        logs.close()
                        pass
                except:
                    logs = open(f'{cwd}/errorlogs.csv','a')
                    print(f'<Неправильная фотография в файле {file} в строке {i-1} у студента {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value}>')
                    logs.write(f'{file}, {i-1}, {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value} \n')
                    logs.close()
                    pass
                ws.cell(row=i, column=1).value = ID #удаление '-' в NCFUGUID
                ws.cell(row=i-1, column=1).value = ws.cell(row=i, column=1).value
                ws.cell(row=i-1, column=2).value = ws.cell(row=i, column=2).value
                ws.cell(row=i-1, column=3).value = ws.cell(row=i, column=3).value
                ws.cell(row=i-1, column=4).value = ws.cell(row=i, column=4).value
                ws.cell(row=i-1, column=5).value = ws.cell(row=i, column=5).value
                ws.cell(row=i-1, column=6).value = ws.cell(row=i, column=6).value
                ws.cell(row=i-1, column=7).value = ws.cell(row=i, column=7).value
                i += 1
            ws._images = [] # удаление всех фото из файла
            filename = file.split("\\")[-1]
            wb.save(filename = f'{cwd}/готовое/{filename}')
            wb.close()
            print(f'{file} успех')
        except:
            messagebox.showerror(title='Ошибка', message='Что-то пошло не так')
    messagebox.showinfo(title='Успех', message='Файлы готовы')

def after_print_student():
    photofold = get_photodirectory() + '/'
    cwd = get_directory()
    files = [f for f in glob.glob(cwd + '**/*.xlsx', recursive=True)]
    os.mkdir(f'{cwd}/готовое')
    os.mkdir(f'{cwd}/готовое/csv')
    bar = Progressbar(root, length=200, style='black.Horizontal.TProgressbar')
    bar.grid(row=7,column=1,stick='we')
    bars = 100 / len(files)
    id = 0
    for file in files:
        bar['value'] += bars
        id += 1
        wb = openpyxl.load_workbook(file)
        ws = wb.worksheets[0]
        ws.insert_rows(1,1)
        i = 2
        row = ws.max_row + 1
        ws.cell(row=1, column=1).value = 'Personnel ID'
        ws.cell(row=1, column=2).value = 'Last Name'
        ws.cell(row=1, column=3).value = 'First Name'
        ws.cell(row=1, column=4).value = 'Middle Name'
        ws.cell(row=1, column=7).value = 'Photo'
        ws.cell(row=1, column=8).value = 'Personnel Type'
        ws.cell(row=1, column=9).value = 'Card #'
        ws.cell(row=1, column=10).value = 'Access Level'
        ws.cell(row=1, column=11).value = 'Badge Format'
        ws.cell(row=1, column=12).value = 'Badge Type'
        ws.cell(row=1, column=13).value = 'Watch level'
        try:
            while i != row:
                photo = ws.cell(row=i, column=7).value
                try:
                    cardid = ws.cell(row=i, column=9).value
                    ws.cell(row=i, column=9).value = int(cardid)
                except:
                    log = open(f'{cwd}/logs.txt', 'a')
                    log.write(f'Отсутствует номер карты в файле {file} в строке {i-1}\n')
                    log.close()
                ws.cell(row=i, column=7).value = photofold + photo
                ws.cell(row=i, column=8).value = 'Student'
                ws.cell(row=i, column=10).value = 'All turnstile'
                ws.cell(row=i, column=11).value = '34bit_noFAC'
                ws.cell(row=i, column=12).value = 'Standard'
                ws.cell(row=i, column=13).value = 'Студенты'
                i += 1
            wb.save(filename = f'{cwd}/готовое/{id}.xlsx')
            filexslx = f'{cwd}/готовое/{id}.xlsx'
            data_xls = pd.read_excel(filexslx, dtype=str, index_col=None)
            data_xls.to_csv(f'{cwd}/готовое/csv/{id}.csv', encoding='cp1251', index=False)
        except:
            print(f'Какаято проблема в {file} в строке {i}')
            pass
    messagebox.showinfo(title='Успех', message='Файлы готовы')

def after_print_employee():
    photofold = get_photodirectory() + '/'
    cwd = get_directory()
    files = [f for f in glob.glob(cwd + '**/*.xlsx', recursive=True)]
    os.mkdir(f'{cwd}/готовое')
    os.mkdir(f'{cwd}/готовое/csv')
    bar = Progressbar(root, length=200, style='black.Horizontal.TProgressbar')
    bar.grid(row=7,column=1,stick='we')
    bars = 100 / len(files)
    id = 0
    for file in files:
        bar['value'] += bars
        id += 1
        wb = openpyxl.load_workbook(file)
        ws = wb.worksheets[0]
        ws.insert_rows(1,1)
        i = 2
        row = ws.max_row + 1
        ws.cell(row=1, column=1).value = 'Personnel ID'
        ws.cell(row=1, column=2).value = 'Last Name'
        ws.cell(row=1, column=3).value = 'First Name'
        ws.cell(row=1, column=4).value = 'Middle Name'
        ws.cell(row=1, column=5).value = 'Photo'
        ws.cell(row=1, column=6).value = 'Personnel Type'
        ws.cell(row=1, column=8).value = 'Card #'
        ws.cell(row=1, column=9).value = 'Access Level'
        ws.cell(row=1, column=10).value = 'Badge Format'
        ws.cell(row=1, column=11).value = 'Badge Type'
        ws.cell(row=1, column=12).value = 'Watch level'
        try:
            while i != row:
                photo = ws.cell(row=i, column=5).value
                try:
                    cardid = ws.cell(row=i, column=8).value
                    ws.cell(row=i, column=8).value = int(cardid)
                except:
                    log = open(f'{cwd}/logs.txt', 'a')
                    log.write(f'Отсутствует номер карты в файле {file} в строке {i-1}\n')
                    log.close()
                ws.cell(row=i, column=5).value = photofold + photo
                ws.cell(row=i, column=6).value = 'Employee - Full Time'
                ws.cell(row=i, column=9).value = 'All turnstile'
                ws.cell(row=i, column=10).value = '34bit_noFAC'
                ws.cell(row=i, column=11).value = 'Standard'
                ws.cell(row=i, column=12).value = 'Сотрудники'
                i += 1
            wb.save(filename = f'{cwd}/готовое/{id}.xlsx')
            filexslx = f'{cwd}/готовое/{id}.xlsx'
            data_xls = pd.read_excel(filexslx, dtype=str, index_col=None)
            data_xls.to_csv(f'{cwd}/готовое/csv/{id}.csv', encoding='cp1251', index=False)
        except:
            print(f'Какаято проблема в {file} в строке {i}')
            pass
    messagebox.showinfo(title='Успех', message='Файлы готовы')

def choose_directory(title = 'Выберите директорию'):
    directory = fd.askdirectory(title="Открыть папку", initialdir="/")
    if directory:
        folder.delete(0,'end')
        folder.insert(0,directory)

def choose_photodirectory(title = 'Выберите директорию'):
    directory = fd.askdirectory(title="Открыть папку", initialdir="/")
    if directory:
        photofolder.delete(0,'end')
        photofolder.insert(0,directory)

def get_directory():
    directory = folder.get()
    return directory

def get_photodirectory():
    directory = photofolder.get()
    return directory

def callback(event):
    webbrowser.open_new(event.widget.cget("text"))

tk.Label(root, text='Выберите папку с файлами xlsx').grid(row=0,column=0,stick='w')
tk.Label(root, text='Папка: ').grid(row=1,column=0,stick='w')
folder = tk.Entry(root)
folder.grid(row=1,column=1,stick='we')

btn1 = tk.Button(root, text='Выбрать папку',
                command=choose_directory).grid(row=1,column=2)

tk.Label(root, text='Выберите папку с фотографиями').grid(row=2,column=0,stick='w')
tk.Label(root, text='Папка с фотографиями: ').grid(row=3,column=0,stick='w')
photofolder = tk.Entry(root)
photofolder.grid(row=3,column=1,stick='we')

btn6 = tk.Button(root, text='Выбрать папку',
                command=choose_photodirectory).grid(row=3,column=2)

tk.Label(root, text='Выберите действие').grid(row=4,column=0,stick='w')

btn2 = tk.Button(root, text='Экспорт фотографий студентов',
                command=export_student).grid(row=5,column=0,stick='we')

btn3 = tk.Button(root, text='Экспорт фотографий сотрудников',
                command=export_emloyee).grid(row=5,column=1,stick='we')
btn4 = tk.Button(root, text='После печати студентов',
                command=after_print_student).grid(row=6,column=0,stick='we')

btn5 = tk.Button(root, text='После печати сотрудников',
                command=after_print_employee).grid(row=6,column=1,stick='we')

tk.Label(root, text='Прогресс: ').grid(row=7,column=0,stick='e')


tk.Label(root, text='Разработка: ДИТ СКФУ').grid(row=8,column=0,stick='w')
lbl = tk.Label(root, text=r"https://github.com/CECTPAYTKYMHE", fg="blue", cursor="hand2")
lbl.grid(row=9,column=0,stick='w')
lbl.bind("<Button-1>", callback)

root.mainloop()